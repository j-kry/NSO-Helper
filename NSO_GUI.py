import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import Event, ttk
import openpyxl
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from regex import E

class Person:
    def __init__(self, idNum, name, location, job, tech, device, sim, isSaved):
        self.idNum = idNum
        self.name = name
        self.location = location
        self.job = job
        self.tech = tech
        self.device = device
        self.sim = sim
        self.isSaved = isSaved

    # This will stop fields from filling if you add spaces even if you cast as string
    def __repr__(self):
        return self.name + "" + self.tech + "" + self.device + "" + self.sim

    def getIdNum(self):
        return self.idNum
    def getName(self):
        return self.name
    def getLocation(self):
        return self.location
    def getJob(self):
        return self.job
    def getIsSaved(self):
        return self.isSaved
    def getAllInfo(self):
        return [self.idNum, self.name, self.job, self.location, self.tech, self.device, self.sim]

    def setTech(self, tech):
        self.tech = tech
    def setDevice(self, device):
        self.device = device
    def setSim(self, sim):
        self.sim = sim
    def setIsSaved(self, isSaved):
        self.isSaved = isSaved

def FileOpen():
    #Path = the one from the window
    filePath = askopenfilename(filetypes=[("Microsoft Excel Worksheet", "*.xlsx"), ("All Files", "*.*")])

    #If filePath empty then return
    if not filePath:
        return

    return filePath

def FileSave():
    #Path = the one from the window
    filePath = asksaveasfilename(filetypes=[("Microsoft Excel Worksheet", "*.xlsx"), ("All Files", "*.*")])

    #If filePath empty then return
    if not filePath:
        return

    return filePath

def PopulateFirstListBox():
    # Loop to read worksheet
    # listAllUsers.insert(0, sheet.cell(row=2, column=5).value + " " + sheet.cell(row=2, column=7).value)
    for i in range(numPeople - 1):
        listAllUsers.insert(i, sheet.cell(row=i+2, column=5).value + " " + sheet.cell(row=i+2, column=7).value)

##############################################
#MOVE SELECTED USERS TO NEW LISTBOX -- STEP 2
##############################################
def MoveLists():

    # Bad code is bad but iterates through the selected users
    # and then creates a new Person object for each one
    # and adds that object to the selectedUserInfo list
    # and finally displays the objects name in the new list box

    if(listAllUsers.curselection()):
        i = 0

        for name in listAllUsers.curselection():
            numRow = int(name) + 2
            selectedUserInfo.append(
                                    Person(sheet.cell(row=numRow, column=1).value, 
                                        sheet.cell(row=numRow, column=5).value + " " + sheet.cell(row=numRow, column=7).value, 
                                        sheet.cell(row=numRow, column=14).value + " " + sheet.cell(row=numRow, column=15).value, 
                                        sheet.cell(row=numRow, column=10).value, 
                                        "", 
                                        "", 
                                        "",
                                        False
                                        )
                                    )
            # listSelectedUsers.insert(tk.END, selectedUserInfo[i].getName())

            i+=1

        ###############
        # GO TO STEP 3
        ###############
        SetupNextWindow()

########################
#SECOND WINDOW - STEP 3
########################
def SetupNextWindow():

    listAllUsers.grid_remove()
    btnNext.destroy()
    btnSave.pack(side=tk.RIGHT, padx=10, ipadx=10)
    btnSaveUser.pack(side=tk.RIGHT, padx=10)
    btnDebug.pack(side=tk.RIGHT, padx=10)

    # Create global combo box so it can be accessed in the bind method
    # Select the first item and then run bind method manually to populate the fields
    global menuSelectedUsers
    menuSelectedUsers = ttk.Combobox(frmLists, value=selectedUserInfo)
    menuSelectedUsers.current(0)
    PopulateUserInfo(Event)
    menuSelectedUsers.grid(row=0, column=0, sticky="n")
    menuSelectedUsers.bind("<<ComboboxSelected>>", PopulateUserInfo)


    lblIdNum.grid(row=0, column=0, sticky="e")
    entIdNum.grid(row=0, column=1, sticky="w")
    lblLocation.grid(row=1, column=0, sticky="e")
    entLocation.grid(row=1, column=1, sticky="w")
    lblJob.grid(row=2, column=0, sticky="e")
    entJob.grid(row=2, column=1, sticky="w")
    lblTechs.grid(row=3, column=0, sticky="e")
    menuTechs.grid(row=3, column=1, sticky="w")
    labelDevices.grid(row=4, column=0, sticky="e")
    menuDevices.grid(row=4, column=1, sticky="w")
    lblSIM.grid(row=5, column=0, sticky="e")
    chkSIM.grid(row=5, column=1, sticky="w")

    

# Once again bad code is bad but this is the only way that I could get this to work
# And I don't feel like refining it now that it works
def PopulateUserInfo(e):

    index = menuSelectedUsers.current()

    if(menuSelectedUsers.get() == selectedUserInfo[index].getName()):
        entLocation.config(state=tk.NORMAL)
        entJob.config(state=tk.NORMAL)
        entIdNum.config(state=tk.NORMAL)
        entIdNum.delete(0, tk.END)
        entIdNum.insert(0, selectedUserInfo[index].getIdNum())
        entJob.delete(0, tk.END)
        entJob.insert(0, selectedUserInfo[index].getJob())
        entLocation.delete(0, tk.END)
        entLocation.insert(0, selectedUserInfo[index].getLocation())

        entIdNum.config(state=tk.DISABLED)
        entLocation.config(state=tk.DISABLED)
        entJob.config(state=tk.DISABLED)
    
def SaveUser():
    index = menuSelectedUsers.current()

    selectedUserInfo[index].setDevice(menuDevices.get())
    selectedUserInfo[index].setTech(menuTechs.get())
    selectedUserInfo[index].setSim(checkValue.get())
    selectedUserInfo[index].setIsSaved(True)
    chkSIM.flash()

def PrintUsers():
    for user in selectedUserInfo:
        print(user)

def SaveSpreadsheet():

    unsavedUsers = False

    for user in selectedUserInfo:
        if user.getIsSaved() == False:
            unsavedUsers = True
    
    if(not unsavedUsers):
        # Create new workbook
        createdWb = Workbook()
        # Select the active worksheet
        createdSheet = createdWb.active
        #Titles for top row
        columnTitles = ["ID", "Preferred Name", "Job Title", "Location", "Tech", "Device Type", "SIM"]

        #Loop to set column names
        #Second row to set column widths. Have to pass the column letter in, cannot be a number
        for i in range(0, len(columnTitles)):
            createdSheet.cell(row=1, column=i+1).value = columnTitles[i]
            createdSheet.column_dimensions[sheet.cell(row=1, column=i+1).column_letter].width = 20.0

        #Start entering info on the second row column 1
        rowCount = 2
        colCount = 1

        for users in selectedUserInfo:
            userInfo = users.getAllInfo()
            for i in range(len(columnTitles)):
                createdSheet.cell(row=rowCount, column=colCount).value = userInfo[i]
                colCount += 1
                if(colCount > len(columnTitles)):
                    rowCount += 1
                    colCount = 1

        createdWb.save(FileSave())
    
    else:
        messagebox.showwarning("Warning", "You have unsaved users")



######################
#Begin Main -- STEP 1
######################

# Load the workbook
wb = load_workbook(filename=FileOpen())
# Load the worksheet
sheet = wb.active
# Get the number of rows
numPeople = sheet.max_row

# Declare list to hold People objects
selectedUserInfo = []

techNames = ["Antwon","Craig", "Eric", "Leonard", "Peter"]
deviceNames = ["Elitebook", "MSI", "Thinkbook", "Thinkpad", "X2"]

# Create the window
window = tk.Tk()
window.title("NSO Helper Outer")
window.resizable(width=False, height=False)
window.iconbitmap('mat.ico')
# Present window to the front and keep it at the front
window.attributes('-topmost', 1)
window.eval('tk::PlaceWindow . center')

#Create all widgets
frmLists = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
frmLists.pack(side=tk.LEFT)

frmControls = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
frmControls.pack(side=tk.TOP)

listAllUsers = tk.Listbox(master=frmLists,height=20, width=30, selectmode=tk.MULTIPLE)
listSelectedUsers = tk.Listbox(master=frmLists,height=20, width=30)
valueInsideUsers = tk.StringVar()

lblIdNum = tk.Label(master=frmControls, text="ID:")
entIdNum = tk.Entry(master=frmControls, width=35)

lblLocation = tk.Label(master=frmControls, text="Location:")
entLocation = tk.Entry(master=frmControls, width=35)

lblJob = tk.Label(master=frmControls, text="Job Title:")
entJob = tk.Entry(master=frmControls, width=35)

lblTechs = tk.Label(master=frmControls, text="Tech:")
menuTechs = ttk.Combobox(frmControls, value=techNames)

labelDevices = tk.Label(master=frmControls, text="Device:")
menuDevices = ttk.Combobox(frmControls, value=deviceNames)

lblSIM = tk.Label(master=frmControls, text="SIM:")
checkValue = tk.StringVar()
chkSIM = tk.Checkbutton(master=frmControls, offvalue="no", onvalue="yes", activebackground="green", variable=checkValue)

frmBtns = tk.Frame()

#########################################################
#On next button - Go to MoveLists Method -- GO TO STEP 2
#########################################################
btnNext = tk.Button(master=frmBtns,text="Next", command=MoveLists)
btnSave = tk.Button(master=frmBtns,text="Save Sheet", command=SaveSpreadsheet)
btnSaveUser = tk.Button(master=frmBtns, text="Save this user", command=SaveUser)
btnDebug = tk.Button(master=frmBtns, text="Print Users", command=PrintUsers)

# Place initial widgets in window
listAllUsers.grid(row=0, column=0, sticky="nsew")
frmBtns.pack(side=tk.BOTTOM, fill=tk.X, ipadx=5, ipady=5)
# btnSave.pack(side=tk.RIGHT, padx=10, ipadx=10)
btnNext.pack(side=tk.RIGHT, ipadx=10, padx=10)

# Grab users from the excel sheet and put them in the list box
PopulateFirstListBox()

window.mainloop()